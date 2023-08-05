import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import re
import argparse
import xlsxwriter
from openpyxl.utils import get_column_letter
#empty variables
perc_dict={}
diff_dict={}
#change accordingly
excelbook='class_test.xlsx'

class Standard:
    def get_percentage(self,csvfile):
        df = pd.read_csv(csvfile, sep=',', header='infer',encoding='latin1')
        cols = list(df)
        for col in cols:
            tot = df[col].shape[0]
            tot_count = df[col].count()
            perc = round((tot_count/tot)*100,2)
            tot = str(tot)
            tot_count=str(tot_count)
            perc=str(perc)+'%'
            perc_dict.update({col:perc})
            perc_df = pd.DataFrame(list(perc_dict.items()), columns=['Column', 'Percentage'])
        return perc_df
            
    def get_datatypes(self,csvfile):
        df = pd.read_csv(csvfile, sep=',', header='infer',encoding='latin1')
        dtype_dict = dict(df.dtypes)
        dtype_df = pd.DataFrame(list(dtype_dict.items()), columns=['Column', 'DataType'])
        return dtype_df
    
    def get_difference(self, newfile, oldfile):
        old_df = pd.read_csv(oldfile, sep=',', header='infer',encoding='latin1')
        new_df = pd.read_csv(newfile, sep=',', header='infer',encoding='latin1')
        #getting cols of current and prev df
        old_cols = list(old_df)
        new_cols = list(new_df)
        #checking for cols that are present in the old file and not present in the new file (REMOVED)
        for col in old_cols:
            if col not in new_cols:
                diff_dict.update({col:'REMOVED'})
            if col in new_cols:
                diff_dict.update({col:'NO CHANGE'})
        #checking for cols that are not present in the old file and are added in the new file (ADDED)
        for col in new_cols:
            if col not in old_cols:
                diff_dict.update({col:'ADDED'})        
        diff_df = pd.DataFrame(list(diff_dict.items()), columns=['Column', 'Status'])
        return diff_df
    
    def verify_totals(self,newfile):

        result=''
        verify_dict={}
        lookup_list=[]
        #reading data file into df
        df = pd.read_csv(newfile, sep=',', header='infer',encoding='latin1')
        
        #getting relevant control file name
        newfile_control = newfile.replace("DATA", "CONTROL")
        #reading control file into df
        df2 = pd.read_csv(newfile_control, sep=',', header='infer',encoding='latin1')
        count=0
        #getting relevant lookup col
        for i in range(2,len(df2.columns)):
            lookup_list.append(df2.columns[i])
            lookup_col = df2.columns[i]
            if (df[lookup_col].sum(axis = 0, skipna = True) == df2[lookup_col].sum(axis = 0, skipna = True)):
                count=count+1
            else:
                count=count+0
        #checking all control values with data values
        if (count==(len(df2.columns)-2) and len(df.columns)==df2[df2.columns[0]].sum(axis=0,skipna=True) and len(df.index)==df2[df2.columns[1]].sum(axis=0,skipna=True)):
            result='YES'
        else:
            result='NO'
        #dict with the results
        verify_dict.update({'TOTAL_ROWS':len(df.index), 'TOTAL_COLUMNS':len(df.columns), 'Premiums_Compared':lookup_list, 'ALL MATCH?':result})
        return verify_dict
    
    def check_special_char(self,data):
        regex = re.compile('[^\w\s]|_')
        myset = set()
        special_dict={}
        df = pd.read_csv(data, sep=',', header='infer',encoding='latin1')
        for col in df.columns:
            countx=len(df)
            county=0
            for i in df[col]:
        #if no special characters are found
                if(regex.search(str(i)) == None):
                    countx=countx-1
                else:
                    special = regex.search(str(i)).group()
                    myset.update(special)
                    county=county+1
            if countx != 0 and county>0:
                special_dict.update({col: myset})
                myset = set()
            else:
                special_dict.update({col:'NO SPECIAL CHAR'})
        special_df = pd.DataFrame(list(special_dict.items()), columns=['Column', 'Special'])
        return special_df
     
    def df_to_excel(self, df,newfile):
        book = load_workbook(excelbook)
        writer = pd.ExcelWriter(excelbook, engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, sheet_name = newfile[len(newfile)-10:len(newfile)-4])
        mysheets = writer.sheets
        cur_sheet=mysheets[newfile[len(newfile)-10:len(newfile)-4]]
        #adjusting the cell size
        for num in range(0,25):
            if num<=1:
                pass
            elif num==2:
                i = get_column_letter(num)
                cur_sheet.column_dimensions[i].width = 30
            else:
                i = get_column_letter(num)
                cur_sheet.column_dimensions[i].width = 18
        writer.save()
        writer.close()

    def total_to_excel(self, df,newfile):
        book = load_workbook(excelbook)
        writer = pd.ExcelWriter(excelbook, engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, startcol=7, startrow=0,sheet_name = newfile[len(newfile)-10:len(newfile)-4])
        writer.save()
        writer.close()

class Validate(Standard):
    
    def __init__(self, file1, file2_flag, *args):
        count=0
        list_of_df = []
        file = str(file1)
        #checking if file2 has been passed
        if file2_flag == True:
            old_file = str(args[0])
            for i in range(1,len(args)):
                if args[i] == 'percentage':
                    df1=super().get_percentage(file)
                    list_of_df.append(df1)
                    count+=1
                if args[i] == 'datatypes':
                    df2=super().get_datatypes(file)
                    list_of_df.append(df2)
                    count+=1
                if args[i] == 'difference':
                    df3=super().get_difference(file,old_file)
                    list_of_df.append(df3)
                    count+=1
                if args[i] == 'special':
                    df4=super().check_special_char(file)
                    list_of_df.append(df4)
                    count+=1
                if args[i] == 'total':
                    total_dict=super().verify_totals(file)
                    total_df = pd.DataFrame(list(total_dict.items()), columns=['Checks', 'Results'])
                    super().total_to_excel(total_df,file)
        #if file2 is not passed
        if file2_flag == False:
            for i in args:
                if i == 'percentage':
                    df1=super().get_percentage(file)
                    list_of_df.append(df1)
                    count+=1
                if i == 'datatypes':
                    df2=super().get_datatypes(file)
                    list_of_df.append(df2)
                    count+=1
                if i == 'difference':
                    print('Sorry! Difference needs 2 input files. Only 1 file has been given. Hence, breaking!')
                    break
                if i == 'special':
                    df4=super().check_special_char(file)
                    list_of_df.append(df4)
                    count+=1
                if i == 'total':
                    total_dict=super().verify_totals(file)
                    total_df = pd.DataFrame(list(total_dict.items()), columns=['Checks', 'Results'])
                    super().total_to_excel(total_df,file)
        #merging dfs based on the number of options passed
        if count == 1:
            super().df_to_excel(list_of_df[0],file)
        if count == 2:
            final_df = pd.merge(list_of_df[0],list_of_df[1], on='Column', how='outer')
            super().df_to_excel(final_df,file)
        if count == 3:
            final_df = pd.merge(list_of_df[0],list_of_df[1], on='Column', how='outer')
            final_df = pd.merge(final_df,list_of_df[2], on='Column', how='outer')
            super().df_to_excel(final_df,file)
        if count == 4:
            final_df = pd.merge(list_of_df[0],list_of_df[1], on='Column', how='outer')
            final_df = pd.merge(final_df,list_of_df[2], on='Column', how='outer')
            final_df = pd.merge(final_df,list_of_df[3], on='Column', how='outer')
            super().df_to_excel(final_df,file)

if __name__ == '__main__':
    check_file2 = False
    parser = argparse.ArgumentParser()
    parser.add_argument("-f1","--file1", help='first file', required=True)
    #optional arguments
    parser.add_argument("-f2","--file2", help='second file', required=False)
    parser.add_argument("-op1","--option1", help='Optional Functionality #1', required=False)
    parser.add_argument("-op2","--option2", help='Optional Functionality #2', required=False)
    parser.add_argument("-op3","--option3", help='Optional Functionality #3', required=False)
    parser.add_argument("-op4","--option4", help='Optional Functionality #4', required=False)
    parser.add_argument("-op5","--option5", help='Optional Functionality #5', required=False)
    args = parser.parse_args()
    count=0
    if args.file2:
        check_file2 = True
    if args.option1:
        count+=1
    if args.option2:
        count+=1
    if args.option3:
        count+=1
    if args.option4:
        count+=1
    if args.option5:
        count+=1
    #if file2 is passed - setting flag to True
    if check_file2:
        if count == 5:
            Validate(str(args.file1), True, str(args.file2), str(args.option1), str(args.option2), str(args.option3), str(args.option4), str(args.option5))
        if count == 4:
            Validate(str(args.file1), True, str(args.file2), str(args.option1), str(args.option2), str(args.option3), str(args.option4))
        if count == 3:
            Validate(str(args.file1), True, str(args.file2), str(args.option1), str(args.option2), str(args.option3))
        if count == 2:
            Validate(str(args.file1), True, str(args.file2), str(args.option1), str(args.option2))
        if count == 1:
            Validate(str(args.file1), True, str(args.file2), str(args.option1))
    #if file2 is not passed - setting flag to False
    if check_file2 == False:
        if count == 5:
            Validate(str(args.file1), False, str(args.option1), str(args.option2), str(args.option3), str(args.option4), str(args.option5))
        if count == 4:
            Validate(str(args.file1), False, str(args.option1), str(args.option2), str(args.option3), str(args.option4))
        if count == 3:
            Validate(str(args.file1), False, str(args.option1), str(args.option2), str(args.option3))
        if count == 2:
            Validate(str(args.file1), False, str(args.option1), str(args.option2))
        if count == 1:
            Validate(str(args.file1), False, str(args.option1))