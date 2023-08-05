import pandas as pd
from ta.utils import dropna
from datetime import datetime, timedelta
import ta
import time
import argparse
from multiprocessing import Process, Manager
from openpyxl import load_workbook


'''============================================================================'''
##################################################################################
'''


'''
##################################################################################
'''============================================================================'''


from algos import AlgoRsiBb
from util import log as log
from util import DataSource

def main():
    # TODO : replace the data_path with the home directory
    # TODO : consideer replacing this with a config file
    parser = argparse.ArgumentParser()
    parser.add_argument('--test_mode', default="True", type=str)
    parser.add_argument('--data_source_config', default="paper", type=str)
    parser.add_argument('--data_path', default="C:/Users/Danny/Documents/GitHub/AlgoTrader/data/",type=str)
    parser.add_argument('--algo_type', default="test", type=str)
    args = parser.parse_args()
    data_path = args.data_path
    algo_type = args.algo_type
    
    # Iterate through master csv
    #   skip first line
    # create a thread for each line in the master csv
    #   This will involve creating algo objects for each thread
    #   This will involve creating a thread object
    
    master_wb = load_workbook(filename=data_path + "master.xlsx")
    sheet = master_wb[algo_type]
    sheet_ds = master_wb["DATA_SOURCE"]

    # print(master_wb["info"].cell(column=1,row=1))
    # print(master_wb["range names"])

    return_dict = Manager().dict()
    ds_row = None
    for row in sheet_ds:
        
        if row[0].value == args.data_source_config:
            ds_row = row
            break
            
    if ds_row == None:
        print("no DATA SOURCE found for config: " + str(args.data_source_config))
        exit()

    algo_list = []
    thread_list = []
    i = 1
    while i < int(sheet.max_row):
        i += 1
        if (sheet[i][0].value == None or sheet[i][0].value == ""):                                           # Column with GUID
            break
        if (algo_type != "test" and algo_type not in sheet[i][0].value):  # If not in main.py set algo_type != ""
            continue

        GUID = sheet[i][0].value
        wallet = sheet[i][1].value
        tickers = sheet[i][2].value.split(",")
        tick_period = sheet[i][3].value
        algo_params = sheet[i][4].value.split(",")
        
        data_source = DataSource(key_id=ds_row[1].value ,secret_key=ds_row[2].value ,
                    base_url=ds_row[3].value , mode=ds_row[4].value , tickers=tickers, 
                    start_date=ds_row[5].value ,end_date=ds_row[6].value 
                    )
                    
        algo = AlgoRsiBb(data_path, data_source, GUID, wallet, tickers, tick_period, algo_params)

        if args.test_mode == "False":
            algo.test_mode = False
            data_source.test_mode = False
        else:
            algo.test_mode = True
            data_source.test_mode = True
            
            
        algo_thread = Process(target=algo.run, name=str(i), args=(return_dict,))
        algo_thread.start()
        
        algo_list.append(algo)
        thread_list.append(algo_thread)

    while 0 < len(thread_list):
        remove_list = []
        for thread in thread_list:
            if not thread.is_alive():
                thread.join()
                i = int(thread.name)
                algo_finished = algo_list[i-2]
                ret = return_dict[algo_finished.GUID]
                sheet[i][5].value = ret[0]
                pos = ret[1]
                total_shares = 0
                for shares in list(pos.values()):
                    total_shares += shares
                sheet[i][6].value = total_shares
                share_val = 0
                for p in pos:
                    share_val += pos[p] * algo_finished._get_last_trade(p)
                sheet[i][7].value = share_val
                sheet[i][8].value = share_val + ret[0]
                master_saved = False
                while not master_saved:
                    try:
                        master_wb.save(data_path + "master.xlsx")
                        master_saved = True
                    except Exception as e:
                        print(e)
                        print("CANNOT SAVE TO MASTER")
                remove_list.append(thread)
        for thread in remove_list:
            thread_list.remove(thread)
        time.sleep(1)
    
    
if __name__ == "__main__":
    main()
    exit()


