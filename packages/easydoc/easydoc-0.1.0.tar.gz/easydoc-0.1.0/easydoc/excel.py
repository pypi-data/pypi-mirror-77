#!/usr/bin/env python
# coding: utf-8
# author: Frank YCJ
# email: 1320259466@qq.com
# date: $ $
# desc:
import os
import random
import shutil
import time
import zipfile

from concurrent.futures import ThreadPoolExecutor
from multiprocessing import Pool
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


def write_data_to_multiple_file_proccess(names):
    # ws, data_list, row_start_index, header_column_index
    wb = load_workbook(names[0])
    ws = wb.active
    data_list = names[1]
    row_start_index = names[2]
    header_column_index = names[3]
    file_dir = names[4]
    file_name = names[5]
    index = names[6]

    for row in data_list:
        ws.append(row)

    if file_name:
        file_name = file_name + "-part" + str(index) + ".xlsx"
    else:
        file_name = "part1" + str(index) + ".xlsx"
    full_file_name = file_dir + file_name
    wb.save(full_file_name)



def write_data_to_multiple_file_proccess_style(names):
    # ws, data_list, row_start_index, header_column_index
    wb = load_workbook(names[0])
    ws = wb.active
    data_list = names[1]
    row_start_index = names[2]
    header_column_index = names[3]
    file_dir = names[4]
    file_name = names[5]
    index = names[6]
    content_height = names[7]
    alignment = names[8]
    # header_row_index = names[9]

    row_size = len(data_list)
    for row in range(row_size):
        col_size = len(data_list[row])
        if content_height:
            ws.row_dimensions[row_start_index + row].height = content_height
        for col in range(col_size):
            ws.cell(row=row_start_index + row , column=header_column_index + col).value = data_list[row][col]
            if alignment:
                ws.cell(row=row_start_index + row ,column=header_column_index + col).alignment = alignment

    if file_name:
        file_name = file_name + "-part" + str(index) + ".xlsx"
    else:
        file_name = "part1" + str(index) + ".xlsx"
    full_file_name = file_dir + file_name
    wb.save(full_file_name)

def write_data_to_single_file_proccess(names):
    wb = load_workbook(names[0])
    ws = wb.active
    data_list = names[1]
    row_start_index = names[2]
    header_column_index = names[3]
    file_dir = names[4]
    file_name = names[5]
    index = names[6]
    row_size = len(data_list)

    for row in range(row_size):
        col_size = len(data_list[row])
        for col in range(col_size):
            ws.cell(row=row_start_index + row, column=header_column_index + col).value = data_list[row][col]
    wb.save(names[0])

    # if file_name:
    #     file_name=file_name+"-part"+str(index)+".xlsx"
    # else:
    #     file_name="part1"+str(index)+".xlsx"
    # full_file_name=file_dir+file_name
    # wb.save(full_file_name)


class EasyExcel:
    DEFAULT_DATA_SIZE = 10000

    @classmethod
    def write_data_to_excel(cls, content_list, header_list=[], header_row_index=1, header_column_index=1,
                            header_height=30, content_height=None, file_dir=None, file_name=None, sheet_name=None,
                            column_width_dict={}, merge_value_dict={}):
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active
        # ws = wb.create_sheet("Mysheet")

        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        # 设定单元格的值，三种方式
        # ws.cell(row=1, column=2).value = 99
        # ws.cell(row=1, column=3, value=100)
        # ws['A4'] = 4

        # ws.append(['This is A1', 'This is B1', 'This is C1'])
        # ws.append({'A' : 'This is A1', 'C': 'This is C1'})
        # ws.append({1: 'This is A1', 3 : 'This is C1'})
        full_file_name = cls.generate_file_name(file_dir, file_name)
        for row in range(len(content_list)):
            if content_height:
                ws.row_dimensions[header_row_index + row+1].height = content_height
            ws.append(content_list[row])
        wb.save(full_file_name)
        return full_file_name

    @classmethod
    def write_data_to_excel_style(cls, content_list, header_list=[], header_row_index=1, header_column_index=1,
                                  header_height=30, content_height=None, file_dir=None, file_name=None, sheet_name=None,
                                  column_width_dict={}, merge_value_dict={},
                                  alignment=Alignment(horizontal="left", vertical="center", wrapText=True)):
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)
        for row in range(len(content_list)):
            if content_height:
                ws.row_dimensions[header_row_index + row+1].height = content_height
            for col in range(len(content_list[row])):
                ws.cell(row=header_row_index + row + 1, column=header_column_index + col).value = content_list[row][col]
                if alignment:
                    ws.cell(row=header_row_index + row + 1, column=header_column_index + col).alignment = alignment
        wb.save(full_file_name)
        return full_file_name

    @classmethod
    def get_excel_rows_cols(cls, file_path, sheet_name=None, sheet_index=None):
        wb = load_workbook(file_path)
        ws = None
        if sheet_name:
            ws = wb.get_sheet_by_name(sheet_name)
        elif sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()
        return ws.max_row, ws.max_column

    @classmethod
    def get_excel_cell_value(cls, file_path, row=-1, column=-1, position="A1", sheet_name=None, sheet_index=None):
        wb = load_workbook(file_path)
        ws = None
        if sheet_name:
            ws = wb.get_sheet_by_name(sheet_name)
        elif sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()
        if row != -1 and column != -1:
            return ws.cell(row=row, column=column).value
        else:
            return ws[position]

    @classmethod
    def get_excel_row_value(cls, file_path, row=0, sheet_name=None, sheet_index=None):
        wb = load_workbook(file_path)
        ws = None
        if sheet_name:
            ws = wb.get_sheet_by_name(sheet_name)
        elif sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()
        rows = ws.rows
        return rows[row]

    @classmethod
    def get_excel_column_value(cls, file_path, column=0, sheet_name=None, sheet_index=None):
        wb = load_workbook(file_path)
        ws = None
        if sheet_name:
            ws = wb.get_sheet_by_name(sheet_name)
        elif sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()
        columns = ws.columns
        return columns[column]

    @classmethod
    def get_excel_sheet_names(cls, file_path):
        wb = load_workbook(file_path)
        return wb.get_active_sheet().title

    @classmethod
    def get_excel_active_sheet_name(cls, file_path, sheet_index=None):
        wb = load_workbook(file_path)
        ws = None
        if sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()
        return wb.get_active_sheet().title

    @classmethod
    def read_data_form_excel(cls, file_path, sheet_name=None, sheet_index=None, min_row=None, max_row=None,
                             min_col=None, max_col=None, values_only=False):
        result = []
        wb = load_workbook(file_path)
        if sheet_name:
            ws = wb.get_sheet_by_name(sheet_name)
        elif sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()

        if min_row and max_row and min_col and max_col:
            for row in ws.iter_rows(min_row, max_row, min_col, max_col, values_only):
                line = [cell.value for cell in row]
                result.append(line)
        else:
            for row in ws.rows:
                line = [cell.value for cell in row]
                result.append(line)
        return result

    @classmethod
    def generate_file_name(self, file_path, file_name):
        if file_path and len(file_path) > 2:
            if os.path.exists(file_path):
                shutil.rmtree(file_path)
            os.makedirs(file_path)
        name_time = time.strftime('%Y-%m-%d-%H-%M-%S') + '-' + str(time.time()).split(".")[0] + str(
            random.randrange(999999))
        if file_name is None:
            file_name = ""
        elif file_name != "":
            file_name = file_name + "-"
        excel_name = file_name + name_time + '.xlsx'
        excel_full_name = os.path.join(file_path, excel_name)
        # excel_full_name = file_path+"/"+excel_name
        return excel_full_name

    @classmethod
    def write_data_to_excel_autothreading(cls, content_list, header_list=[], header_row_index=1, header_column_index=1,
                                          header_height=30, file_dir=None, file_name=None,
                                          sheet_name=None,
                                          column_width_dict={}, merge_value_dict={}):
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        thread_count = 1
        content_list_size = len(content_list)
        if content_list_size > cls.DEFAULT_DATA_SIZE:
            if content_list_size % cls.DEFAULT_DATA_SIZE != 0:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE) + 1
            else:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE)
        # pool = ThreadPoolExecutor(thread_count,thread_name_prefix="export_data_to_excel_")
        # pool = ThreadPoolExecutor(max_workers=thread_count, thread_name_prefix="export_data_to_excel_")
        pool = ThreadPoolExecutor(max_workers=thread_count, thread_name_prefix="export_data_to_excel_")
        # pool=threadpool.ThreadPool(thread_count)
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)

        for i in range(thread_count):
            start_account = cls.DEFAULT_DATA_SIZE * i
            if (i + 1) == thread_count:
                pool.submit(cls.write_data_to_file, ws, content_list[start_account:content_list_size],
                            start_account + header_row_index + 1,
                            header_column_index)
            else:
                pool.submit(cls.write_data_to_file, ws,
                            content_list[start_account:(start_account + cls.DEFAULT_DATA_SIZE)],
                            start_account + header_row_index + 1,
                            header_column_index)

        pool.shutdown(wait=True)
        wb.save(full_file_name)

        return full_file_name

    @classmethod
    def write_data_to_excel_autoprocess(cls, content_list, header_list=[], header_row_index=1, header_column_index=1,
                                        header_height=30, file_dir=None, file_name=None,
                                        sheet_name=None,
                                        column_width_dict={}, merge_value_dict={}):
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        thread_count = 1
        content_list_size = len(content_list)
        if content_list_size > cls.DEFAULT_DATA_SIZE:
            if content_list_size % cls.DEFAULT_DATA_SIZE != 0:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE) + 1
            else:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE)
        pool = Pool(thread_count)
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)
        wb.save(full_file_name)
        # args=[]
        for i in range(thread_count):
            start_account = cls.DEFAULT_DATA_SIZE * i
            if (i + 1) == thread_count:
                # args.append([ws, content_list[start_account:content_list_size], start_account + header_row_index+1,header_column_index])
                # pool.map(cls.write_data_to_file_proccess, ws, content_list[start_account:content_list_size], start_account + header_row_index+1,
                #             header_column_index)
                # pool.apply_async(write_data_to_file_proccess, args=([ws, content_list[start_account:content_list_size], start_account + header_row_index+1,header_column_index],))
                pool.apply_async(write_data_to_multiple_file_proccess, args=(
                    [full_file_name, content_list[start_account:content_list_size],
                     start_account + header_row_index + 1,
                     header_column_index, file_dir, file_name, (i + 1)],))
            else:
                # args.append([ws, content_list[start_account:(start_account+cls.DEFAULT_DATA_SIZE)], start_account + header_row_index+1,header_column_index])
                # pool.map(cls.write_data_to_file_proccess, ws, content_list[start_account:(start_account+cls.DEFAULT_DATA_SIZE)], start_account + header_row_index+1,
                #             header_column_index)
                # pool.apply_async(write_data_to_file_proccess, args=([ws, content_list[start_account:(start_account+cls.DEFAULT_DATA_SIZE)], start_account + header_row_index+1,header_column_index],))
                pool.apply_async(write_data_to_multiple_file_proccess, args=(
                    [full_file_name, content_list[start_account:(start_account + cls.DEFAULT_DATA_SIZE)],
                     start_account + header_row_index + 1, header_column_index, file_dir, file_name, (i + 1)],))

        # pool.map(cls.write_data_to_file_proccess,args)
        pool.close()
        pool.join()
        # wb.save(full_file_name)
        if not file_name:
            file_name="data"
        full_file_name=cls.zip_files(file_dir,file_name)

        return full_file_name

    @classmethod
    def write_data_to_file(cls, ws, data_list, row_start_index, header_column_index):
        row_size = len(data_list)
        for row in range(row_size):
            col_size = len(data_list[row])
            for col in range(col_size):
                ws.cell(row=row_start_index + row, column=header_column_index + col).value = data_list[row][col]

        # wb.save(full_file_name)

    @classmethod
    def zip_files(cls, dir_path, new_file_name):
        # file_news = dir_path + "/" + new_file_name + '.zip'  # 压缩后文件夹的名字
        file_path=str(dir_path)
        if file_path.endswith("/"):
            file_path=file_path[:len(file_path)-1]

        if file_path.endswith("\\"):
            file_path = file_path[:len(file_path) - 1]
        file_news = file_path + '.zip'  # 压缩后文件夹的名字

        z = zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED)  # 参数一：文件夹名
        for dirpath, dirnames, filenames in os.walk(dir_path):
            fpath = dirpath.replace(dir_path, '')  # 这一句很重要，不replace的话，就从根目录开始复制
            fpath = fpath and fpath + os.sep or ''  # 实现当前文件夹以及包含的所有文件的压缩
            for filename in filenames:
                z.write(os.path.join(dirpath, filename), fpath + filename)
        z.close()
        return file_news

    @classmethod
    def write_data_to_excel_autoprocess_style(cls, content_list, header_list=[], header_row_index=1,
                                                header_column_index=1,
                                                header_height=30, content_height=None, file_dir=None, file_name=None,
                                                sheet_name=None,
                                                column_width_dict={}, merge_value_dict={},
                                                alignment=Alignment(horizontal="center", vertical="center",
                                                                    wrapText=True)):
        thread_count = 1
        content_list_size = len(content_list)
        if content_list_size > cls.DEFAULT_DATA_SIZE:
            if content_list_size % cls.DEFAULT_DATA_SIZE != 0:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE) + 1
            else:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE)
        pool = Pool(thread_count)
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)
        wb.save(full_file_name)
        for i in range(thread_count):
            start_account = cls.DEFAULT_DATA_SIZE * i
            if (i + 1) == thread_count:
                pool.apply_async(write_data_to_multiple_file_proccess_style, args=(
                    [full_file_name, content_list[start_account:content_list_size],
                     header_row_index + 1,
                     header_column_index, file_dir, file_name, (i + 1),content_height,alignment],))
            else:
                pool.apply_async(write_data_to_multiple_file_proccess_style, args=(
                    [full_file_name, content_list[start_account:(start_account + cls.DEFAULT_DATA_SIZE)],
                     header_row_index + 1, header_column_index, file_dir, file_name, (i + 1),content_height,alignment],))

        pool.close()
        pool.join()
        # wb.save(full_file_name)
        if not file_name:
            file_name = "data"
        full_file_name = cls.zip_files(file_dir, file_name)

        return full_file_name


    @classmethod
    def write_data_to_excel_autothreading_style(cls, content_list, header_list=[], header_row_index=1,
                                                header_column_index=1,
                                                header_height=30, content_height=None, file_dir=None, file_name=None,
                                                sheet_name=None,
                                                column_width_dict={}, merge_value_dict={},
                                                alignment=Alignment(horizontal="center", vertical="center",
                                                                    wrapText=True)):
        thread_count = 1
        content_list_size = len(content_list)
        if content_list_size > cls.DEFAULT_DATA_SIZE:
            if content_list_size % cls.DEFAULT_DATA_SIZE != 0:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE) + 1
            else:
                thread_count = (content_list_size / cls.DEFAULT_DATA_SIZE)
        # pool = ThreadPoolExecutor(thread_count,thread_name_prefix="export_data_to_excel_")
        pool = ThreadPoolExecutor(max_workers=thread_count, thread_name_prefix="export_data_to_excel_")
        # pool=threadpool.ThreadPool(thread_count)
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)

        for i in range(thread_count):
            start_account = cls.DEFAULT_DATA_SIZE * i
            if (i + 1) == thread_count:
                pool.submit(cls.write_data_to_file_style, ws,
                            content_list[start_account:content_list_size], start_account + header_row_index+1,
                            header_column_index, content_height, alignment)
            else:
                pool.submit(cls.write_data_to_file_style, ws,
                            content_list[start_account:(start_account + cls.DEFAULT_DATA_SIZE)],
                            start_account + header_row_index+1, header_column_index, content_height,
                            alignment)

        pool.shutdown(wait=True)
        wb.save(full_file_name)

        return full_file_name

    @classmethod
    def write_data_to_file_style(cls, ws, data_list, row_start_index, header_column_index,
                                 content_height,alignment):

        row_size = len(data_list)
        for row in range(row_size):
            col_size = len(data_list[row])
            if content_height:
                ws.row_dimensions[row_start_index + row].height = content_height
            for col in range(col_size):
                ws.cell(row=row_start_index + row , column=header_column_index + col).value = data_list[row][col]
                if alignment:
                    ws.cell(row=row_start_index + row ,
                            column=header_column_index + col).alignment = alignment
        # wb.save(full_file_name)

    @classmethod
    def write_data_to_excel_multithreading(cls, content_list, thread_count, header_list=[], header_row_index=1,
                                           header_column_index=1,
                                           header_height=30, content_height=None, file_dir=None, file_name=None,
                                           sheet_name=None,
                                           column_width_dict={}, merge_value_dict={}):
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        content_list_size = len(content_list)
        pool = ThreadPoolExecutor(max_workers=thread_count, thread_name_prefix="export_data_to_excel_")
        # pool=threadpool.ThreadPool(thread_count)
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)

        splice_size = content_list_size / thread_count

        for i in range(thread_count):
            start_account = splice_size * i
            if (i + 1) == thread_count:
                pool.submit(cls.write_data_to_file, ws, content_list[start_account:content_list_size],
                            start_account + header_row_index + 1,
                            header_column_index)
            else:
                pool.submit(cls.write_data_to_file, ws, content_list[start_account:(start_account + splice_size)],
                            start_account + header_row_index + 1,
                            header_column_index)

        pool.shutdown(wait=True)
        wb.save(full_file_name)

        return full_file_name

    @classmethod
    def write_data_to_excel_multithreading_style(cls, content_list, thread_count, header_list=[], header_row_index=1,
                                                 header_column_index=1,
                                                 header_height=30, content_height=None, file_dir=None, file_name=None,
                                                 sheet_name=None,
                                                 column_width_dict={}, merge_value_dict={},
                                                 alignment=Alignment(horizontal="center", vertical="center",
                                                                     wrapText=True)):
        content_list_size = len(content_list)
        pool = ThreadPoolExecutor(max_workers=thread_count, thread_name_prefix="export_data_to_excel_")
        # pool=threadpool.ThreadPool(thread_count)
        wb = Workbook()
        ws = None
        if sheet_name:
            ws = wb.add_sheet(sheet_name)
        else:
            ws = wb.active

        if header_list or merge_value_dict:
            ws.row_dimensions[header_row_index].height = header_height
        for col in range(len(header_list)):
            ws.cell(row=header_row_index, column=col + header_column_index).value = header_list[col]
            ws.cell(row=header_row_index, column=col + header_column_index).alignment = alignment
        for name, width in column_width_dict.items():
            ws.column_dimensions[name].width = width

        for cell, value in merge_value_dict.items():
            ws.merge_cells(cell)
            ws[cell.split(":")[0]] = value

        full_file_name = cls.generate_file_name(file_dir, file_name)

        splice_size = content_list_size / thread_count

        for i in range(thread_count):
            start_account = splice_size * i
            if (i + 1) == thread_count:
                pool.submit(cls.write_data_to_file_style, ws,
                            content_list[start_account:content_list_size], start_account + header_row_index+1,
                            header_column_index, content_height, alignment)
            else:
                pool.submit(cls.write_data_to_file_style, ws, content_list[start_account:(start_account + splice_size)],
                            start_account + header_row_index+1,  header_column_index, content_height,
                            alignment)

        pool.shutdown(wait=True)
        wb.save(full_file_name)

        return full_file_name

    @classmethod
    def get_object_verbose_name_for_header(cls, object, include_keys_list=None, exclude_keys_list=None):
        # keys_dict={}    # 由于dict的无序性，考虑用双list解决
        keys_list = []
        value_list = []
        for f in object._meta.fields:
            try:
                name = f.name
                value = str(f.verbose_name)
            except:
                value = ""
            # keys_dict[name] = value
            keys_list.append(name)
            value_list.append(value)
        header_list = []
        if include_keys_list:
            for key in include_keys_list:
                # header_list.append(keys_dict[key])
                header_list.append(value_list[keys_list.index(key)])

        elif exclude_keys_list:
            # for k,v in keys_dict.items():
            for k in keys_list:
                if not exclude_keys_list.__contains__(k):
                    # header_list.append(k)
                    header_list.append(value_list[keys_list.index(k)])
        else:
            # for k,v in keys_dict.items():
            #     header_list.append(v)
            header_list = value_list
        return header_list

    @classmethod
    def write_data_to_excel_by_file(cls, file_path, content_list, header_row_index=1, header_column_index=1,
                                    file_dir=None, file_name=None, sheet_index=None, sheet_name=None):
        wb = load_workbook(file_path)
        if sheet_name:
            ws = wb.get_sheet_by_name(sheet_name)
        elif sheet_index:
            sheet_names = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheet_names[sheet_index])
        else:
            ws = wb.active  # 等同于 ws = wb.get_active_sheet()

        for row in range(len(content_list)):
            for col in range(len(content_list[row])):
                ws.cell(row=row + header_row_index, column=header_column_index + col).value = content_list[row][col]

        full_file_name = cls.generate_file_name(file_dir, file_name)
        wb.save(full_file_name)
        return full_file_name
