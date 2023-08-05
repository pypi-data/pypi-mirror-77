#!/usr/bin/env python
# -*- coding: utf-8 -*-

from typing import *
import openpyxl
from pathlib import Path
from cytoolz.curried import pipe, take, map, curry, compose, merge, filter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import re
from collections import Counter
from dataclasses import dataclass, field, asdict, astuple
from tabulate import tabulate
import atek
from fnmatch import fnmatch


def workbook(path: Path) -> openpyxl.Workbook:
    return pipe(
        path
        ,atek.valid_path
        ,str
        ,openpyxl.load_workbook
    )


@dataclass(frozen=True)
class Format:
    header: str
    number_format: str = "General"


    @classmethod
    def from_tuple(cls, tup: Tuple[str, str]):
        return cls(*tup)


    def astuple(self) -> Tuple[str, str]:
        return astuple(self)


    def asdict(self) -> Dict[str,str]:
        return asdict(self)


@dataclass(frozen=True)
class Column:
    header: str
    header_range: str
    data_range: str
    data_type: str
    max_width: int


    @classmethod
    def from_cells(cls, sheet_name: str, cells: List[openpyxl.cell.cell.Cell]):
        return cls(
            header       = cells[0].value or "",
            header_range = cls.format_range(sheet_name, [cells[0]]),
            data_range   = cls.format_range(sheet_name, cells[1:]),
            data_type    = cls.dtypes(cells),
            max_width    = cls.max_len(cells),
        )


    @staticmethod
    def max_len(cells: List[Cell]) -> int:
        return max([
            len(str(cell.value))
            for cell in cells
        ])


    @staticmethod
    def dtypes(cells: List[Cell]) -> str:
        data_types = Counter([
            cell.data_type
            for cell in cells[1:]
        ])

        most_common = data_types.most_common(1)[0][0]
        mapping = {
            "s": "text",
            "n": "number",
            "d": "date",
            "f": "formula",
        }

        return mapping.get(most_common, most_common)


    @staticmethod
    def format_range(cells: List[Cell]) -> str:
        return (
            f"{cells[0].coordinate}" if len(cells) == 1 else
            f"{cells[0].coordinate}:{cells[-1].coordinate}"
        )


@dataclass(frozen=True)
class ColumnFormat:
    column: Column
    number_format: str


@dataclass(frozen=True)
class Sheet:
    sheet_name: str
    sheet_range: str
    columns: List[Column]


    @classmethod
    def from_wb(cls, wb: Workbook, sheet_name: str):
        ws = wb[sheet_name]
        sheet_range = ws.dimensions
        columns = [
            Column.from_cells(sheet_name, cells)
            for cells in ws.iter_cols()
        ]
        return cls(sheet_name, sheet_range, columns)


    @classmethod
    def from_path(cls, path: Path, sheet_name: str):
        wb = workbook(path)
        return cls.from_wb(wb, sheet_name)


    def as_records(self) -> List[Dict[str,Any]]:
        sheet = {
            k: v for k, v in asdict(self).items()
            if k != "columns"
        }
        return [merge(sheet, asdict(col)) for col in self.columns]


    def cols_asdict(self):
        return {col.header: col for col in self.columns}


def unix_match(sheet: Sheet, *formats: Format) -> Iterable[ColumnFormat]:
    matched = set()
    for col in sheet.columns:
        # If match yield Format object
        for fmt in formats:
            if col.header in matched:
                pass
            if fnmatch(col.header, fmt.header):
                yield ColumnFormat(col, fmt.number_format)
                matched.add(col.header)
        if col.header not in matched:
            # If unmatched yield Format object with default formats
            yield ColumnFormat(col, "General")
            matched.add(col.header)


@curry
def sheet_info(sheet_name: str, path: Path) -> Sheet:
    return Sheet.from_path(path, sheet_name)


@curry
def format_sheet(
    sheet_name: str,
    path: Path,
    *formats: Tuple[str, str]) -> Path:

    wb = workbook(path)
    ws = wb[sheet_name]
    for cells in ws.iter_cols():
        header = cells[0].value
        data = cells[1:]
        for fmt in formats:
            if fnmatch(header, fmt.header):
                # Apply header formats

                # Apply data formats
                for cell in data:
                    cell.number_format = fmt.number_format
    save_path = pipe(path, atek.valid_path, str)
    wb.save(save_path)
    wb.close()


pipe(
    atek.config("ATEK_LOCAL", "paths", "reports")
    ,lambda path: Path(path).expanduser() / "Turn Time/Turntimes 2020-08-18.xlsx"
    ,sheet_info("Turn Times")
    ,lambda sheet: unix_match(sheet, Format("*Orders", "#,##0"), Format("*Median*", "#,##0.00"), Format("*Q[13]*", "#,##0.00"))
    ,map(asdict)
    ,atek.show(column_width=40)
)


@curry
def table_name(sheet_name: str) -> str:
    new_name = re.sub("[\\\/\*\[\]:\? ]", "_", sheet_name)
    return new_name


def add_table(path: Path, sheet_naem: str, display_name: str="",
    showFirstColumn: bool=False, showLastColumn=False,
    showColumnStripes: bool=False, showRowStripes: bool=True,
    style_name: str="TableStyleLight15") -> Path:

    wb = openpyxl.load_workbook(str(path.expanduser()))
    ws = wb[sheet_name]

    data_range = ws.dimensions
    display_name = display_name or table_name(ws.title)

    table = Table(displayName=display_name, ref=data_range)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
        showRowStripes=True,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(str(path))

    return path
