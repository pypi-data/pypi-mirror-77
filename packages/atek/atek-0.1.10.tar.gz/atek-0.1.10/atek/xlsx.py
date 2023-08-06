#!/usr/bin/env python
# -*- coding: utf-8 -*-

import typing as tp
import openpyxl
from pathlib import Path
import cytoolz.curried as tz
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
import re
from collections import Counter
from dataclasses import dataclass, field, asdict, astuple
import atek
from fnmatch import fnmatch

__all__ = "fcell format_sheet add_table format_file".split()

def workbook(path: Path) -> Workbook:
    """Returns an Workbook object"""
    return tz.pipe(
        path
        ,atek.valid_path
        ,str
        ,openpyxl.load_workbook
    )


@dataclass(frozen=True)
class CellFormat:
    font: Font
    align: Alignment
    number_format: str
    column_width: int
    

class Format(tp.NamedTuple):
    pattern: str
    cell_format: CellFormat

    def validate(self):
        assert isinstance(self.cell_format, CellFormat), \
            "cell_format is not of type CellFormat"
        assert isinstance(self.pattern, str), "pattern must be a string"


def fcell(**kwargs) -> CellFormat:
    """Creates default formats and merges them with provided formats.
    Used as a helper function to format_sheet."""
    # Set base format 
    base = {"horizontal": "left", "vertical": "top", "wrap_text": True,
        "column_width": 20, "dtype": "", "number_format": ""}

    # Merge in provided args
    kwargs = kwargs or {}
    all_args = tz.merge(base, kwargs)

    # Calculate the dtype if not provided
    dtype = all_args["dtype"]
    num = all_args["number_format"].lower()
    right = {"horizontal": "right"}
    comma2 = {"number_format": "#,##0.00"}
    ymd = {"number_format": "yyyy-mm-dd"}
    extra = (
        right if dtype in ["n", "d"] and num != "" else
        tz.merge(right, comma2) if dtype == "n" and num == "" else
        tz.merge(right, ymd) if dtype == "d" and num == "" else
        right if "#" in num else
        right if "%" in num else
        right if "yy" in num else
        right if "m" in num else
        right if "d" in num else
        {} 
    )

    all_args = tz.merge(all_args, extra)

    font_args = (
        "name bold italic strike outline shadow condense "
        "color extend size underline vertAlign"
    ).split()

    align_args = (
        "horizontal vertical text_rotation wrap_text shrink_to_fit "
        "indent"
    ).split()

    get_args = lambda arg_set: {
        k: v 
        for k, v in all_args.items() 
        if k in arg_set
    }

    return CellFormat(
        font=Font(**get_args(font_args)),
        align=Alignment(**get_args(align_args)),
        number_format=all_args["number_format"],
        column_width=all_args["column_width"]
    )


def match(dtype: str, column_name: str, *formats: Format) -> CellFormat:
    for fmt in formats:
        # If normal tuple is provided then convert to Format
        _fmt = fmt if isinstance(fmt, Format) else Format(*fmt)
        _fmt.validate()

        # If match found return value
        if fnmatch(column_name, _fmt.pattern):
            return _fmt.cell_format

    # If no matches found, return default Format
    return fcell(dtype=dtype)


@tz.curry
def format_sheet(
    sheet_name: str,
    path: Path,
    *formats: Format,
    freeze_panes: str="") -> Path:

    wb = workbook(path)
    ws = wb[sheet_name]

    #  for column_index, cells in enumerate(ws.iter_cols(), start=1):
    for cells in ws.iter_cols():

        header = cells[0]
        column_name = header.value
        data = cells[1:]
        data_types = Counter(cell.data_type for cell in data)
        column_dtype = data_types.most_common(1)[0][0]
        max_width = max(len(str(cell.value)) for cell in cells)

        # Set freeze_panes for the first cell if it was not provided
        if freeze_panes == "":
            freeze_panes = data[0].coordinate

        # Get formats
        hfmt = fcell(dtype=column_dtype)
        dfmt = match(column_dtype, column_name, *formats)

        # Set column width
        width = min(max_width + 3, dfmt.column_width)
        ws.column_dimensions[header.column_letter].width = width
    #
        # Apply header format
        header.number_format = hfmt.number_format
        header.font = hfmt.font
        header.alignment = hfmt.align

        # Apply data formats
        for cell in data:
            cell.number_format = dfmt.number_format
            cell.font = dfmt.font
            cell.alignment = dfmt.align

    #  # Set the freeze_panes
    ws.freeze_panes = freeze_panes

    save_path = tz.pipe(path, atek.valid_path, str)
    wb.save(save_path)

    return save_path


@tz.curry
def table_name(sheet_name: str) -> str:
    new_name = re.sub("[\\\/\*\[\]:\? ]", "_", sheet_name)
    return new_name


@tz.curry
def add_table(sheet_name: str, path: Path, display_name: str="",
    showFirstColumn: bool=False, showLastColumn=False,
    showColumnStripes: bool=False, showRowStripes: bool=True,
    style_name: str="TableStyleLight15") -> Path:

    wb = workbook(path)
    ws = wb[sheet_name]

    data_range = ws.dimensions
    display_name = display_name or table_name(ws.title)

    table = Table(displayName=display_name, ref=data_range)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
        showRowStripes=True,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    save_path = tz.pipe(path, atek.valid_path, str)
    wb.save(save_path)

    return save_path


def format_file(path: Path) -> Path:
    """Formats all sheets in a file using the default formats and add
    worksheet tables to each sheet."""
    wb = workbook(path)
    for sheet in wb.sheetnames:
        tz.pipe(
            path
            ,format_sheet(sheet)
            ,add_table(sheet)
        )

    return atek.valid_path(path)


if __name__ == "__main__":
    tz.pipe(
        atek.config("ATEK_LOCAL", "paths", "reports")
        ,lambda path: Path(path).expanduser() / "Turn Time/Turntimes 2020-08-20.xlsx"
        ,format_file
        ,print
        #  ,lambda path: format_sheet("State Data", path,
            #  Format("*Orders", fcell(number_format="#,##0")),
        #  )
        #  ,add_table("State Data")
    )
