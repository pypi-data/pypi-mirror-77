#!/usr/bin/env python
# -*- coding: utf-8 -*-

from typing import *
import openpyxl
from pathlib import Path
from atek.connect import paths, print_dict
from cytoolz.curried import pipe, take, map, curry
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
from collections import Counter 
from dataclasses import dataclass, field
from tabulate import tabulate

print_dict("paths", paths())


def sheet_data_range(path: Path):
    wb = openpyxl.load_workbook(str(path))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"{ws.title=}, {ws.dimensions=}")


def to_list(values: Union[str, Iterable[str]]) -> List[str]:
    if isinstance(values, str):
        return [values]
    else:
        return [v for v in values]


@dataclass(frozen=True)
class Column:
    header: str
    header_range: str 
    data_range: str 
    data_type: str 
    max_width: int 

    
    @classmethod
    def from_cells(cls, sheet_name: str, cells: List[openpyxl.cell.cell.Cell]): 
        def convert_data_type(counter: Counter) -> str:
            most_common = counter.most_common(1)[0][0]
            mapping = {
                "s": "text",
                "n": "number",
                "d": "date",
                "f": "formula",
            }
            return mapping.get(most_common, most_common)

        def format_range(sheet, rng):
            return f"'{sheet}'!{rng}"

        header = cells[0]

        data_range = f"{cells[1].coordinate}:{cells[-1].coordinate}"

        data_type = Counter([
            cell.data_type 
            for cell in cells[1:]
        ])

        max_len = max([
            len(str(cell.value)) 
            for cell in cells
        ])

        return cls(
            header.value,
            format_range(sheet_name, header.coordinate),
            format_range(sheet_name, data_range),
            convert_data_type(data_type),
            max_len,
        )


@dataclass(frozen=True)
class Sheet:
    sheet_name: str 
    sheet_range: str
    columns: Tuple[Column]


    @classmethod
    def from_wb(cls, wb: openpyxl.workbook.Workbook, sheet_name: str):
        ws = wb[sheet_name]
        sheet_range = ws.dimensions
        columns = [
            Column.from_cells(sheet_name, cells)
            for cells in ws.iter_cols()
        ]
        return cls(sheet_name, sheet_range, columns)


@dataclass(frozen=True)
class xlsx:
    sheets: Tuple[Sheet] 
    path: Path


    @classmethod
    def from_path(cls, path: Union[str, Path]): 
        wb = openpyxl.load_workbook(str(path))
        sheets = [
            Sheet.from_wb(wb, sheet_name)
            for sheet_name in wb.sheetnames 
        ]
        return cls(sheets, path)


    def __str__(self):
        text = f"{self.path=}"
        for sheet in self.sheets:
            text += (
                f"\n\tsheet_name  = {sheet.sheet_name}"
                f"\n\tsheet_range = {sheet.sheet_range}"
            )
            for col in sheet.columns:
                text += f"\n\t\t{col}"
        return text


pipe(
    paths()["reports"]
    ,lambda p: Path(p).expanduser().glob("*.xlsx")
    ,map(xlsx.from_path)
    ,lambda files: [print(file) for file in files]
)
            

def sheet_names(path: Path) -> List[Sheet]:
    wb = openpyxl.load_workbook(str(path))
    return [
        Sheet(path, sheet)
        for sheet in wb.sheetnames
    ]


@curry
def column_ranges(sheet_name: str, path: Path):

    def format_range(sheet, rng):
        return f"'{sheet}'!{rng}"

    wb = openpyxl.load_workbook(str(path))
    ws = wb[sheet]
    columns = {}
    for cells in ws.iter_cols():
        ranges = {}

        header = cells[0]

        data_range = f"{cells[1].coordinate}:{cells[-1].coordinate}"

        data_type = Counter([
            cell.data_type 
            for cell in cells[1:]
        ]).most_common(1)[0][0]

        max_len = max([
            len(str(cell.value)) 
            for cell in cells
        ])

        ranges = {
            "column_letter": header.column_letter,
            "column_range": format_range(sheet, f"{header.column_letter}:{header.column_letter}"),
            "header_range": format_range(sheet, header.coordinate),
            "data_range": format_range(sheet, data_range),
            "data_type": {"s": "text", "n": "number", "d": "date"}[data_type],
            "max_width": max_len,
        }

        columns[header.value] = ranges 

    return columns


@curry
def table_name(sheet_name: str) -> str:
    new_name = re.sub("[\\\/\*\[\]:\? ]", "_", sheet_name)
    return new_name


def add_table(path: Path, sheet_naem: str, display_name: str="",
    showFirstColumn: bool=False, showLastColumn=False,
    showColumnStripes: bool=False, showRowStripes: bool=True,
    style_name: str="TableStyleLight15") -> Path:

    wb = openpyxl.load_workbook(str(path.expanduser()))
    ws = wb[sheet_name]

    data_range = ws.dimensions
    display_name = display_name or table_name(ws.title)

    table = Table(displayName=display_name, ref=data_range)
    style = TableStyleInfo(
        name=style_name, 
        showFirstColumn=False, 
        showLastColumn=False, 
        showColumnStripes=False,
        showRowStripes=True,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(str(path))

    return path


#  pipe(
    #  Path(paths()["mercury"]).expanduser().glob("*.xlsx")
    #  ,take(1)
    #  ,map(sheet_names)
    #  ,map(lambda d: [
        #  column_ranges(sheet, d["path"])
        #  for sheet in d["sheets"]
    #  ])
    #  ,map(lambda d: [
        #  print_dict(f"{sheet}!{key}", value)
        #  for sheet, column in d.items()
        #  for key, value in column.items()
    #  ])
    #  ,map(lambda files: [
        #  print_dict("sheet_names", file)
        #  for file in files
    #  ])
    #  ,list
    #  ,print
#  )
